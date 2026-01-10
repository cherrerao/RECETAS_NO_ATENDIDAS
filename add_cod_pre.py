import shutil
from pathlib import Path

from openpyxl import load_workbook


def main():
    folder = Path(__file__).parent
    xlsx = folder / "catalogo-redes.xlsx"
    if not xlsx.exists():
        print(f"Archivo no encontrado: {xlsx}")
        return

    backup = folder / (xlsx.stem + ".backup" + xlsx.suffix)
    shutil.copyfile(xlsx, backup)
    print(f"Backup creado: {backup.name}")

    wb = load_workbook(xlsx)
    ws = wb.active

    # Buscar la columna del encabezado 'COD PR' (busca coincidencia parcial)
    header_row = 1
    cod_pr_col = None
    for cell in ws[header_row]:
        if cell.value and isinstance(cell.value, str) and 'COD PR' in cell.value.upper():
            cod_pr_col = cell.column
            break

    if cod_pr_col is None:
        print("No se encontró la columna 'COD PR' en la primera fila.")
        return

    # Nuevo índice: inmediatamente después de COD PR
    new_col_idx = cod_pr_col + 1
    ws.insert_cols(new_col_idx)
    ws.cell(row=header_row, column=new_col_idx, value='COD PRE')

    for r in range(header_row + 1, ws.max_row + 1):
        ws.cell(row=r, column=new_col_idx).value = ws.cell(row=r, column=cod_pr_col).value

    wb.save(xlsx)
    print(f"Columna 'COD PRE' agregada y poblada en {xlsx.name}")


if __name__ == '__main__':
    main()
